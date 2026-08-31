import unittest
from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook

from excel_export import read_schedule_workbook, update_schedule_workbook
from scheduler import Assignment, MANAGER


class ExcelExportTests(unittest.TestCase):
    def test_updates_only_primary_shift_cells(self):
        book = Workbook()
        sheet = book.active
        sheet.title = "26.09"
        sheet["B1"] = "9/1"
        sheet["A2"] = "front 10-19"
        sheet["A3"] = "front help"
        sheet["B3"] = "Keep me"
        sheet["A4"] = "memo"
        sheet["B4"] = "Keep memo"
        sheet["A5"] = "Clean A"
        sheet["A6"] = "Clean B"
        sheet["A7"] = "clean help"
        sheet["B7"] = "Keep helper"
        sheet["A8"] = "宿直(19-10)"
        source = BytesIO()
        book.save(source)

        day = date(2026, 9, 1)
        result = update_schedule_workbook(
            source,
            [
                Assignment(day, "Front", "Alex"),
                Assignment(day, "Cleaning", "Jamie"),
                Assignment(day, "Cleaning", "Morgan"),
                Assignment(day, "Night", "Taylor"),
            ],
            day,
        )
        updated = load_workbook(BytesIO(result))["26.09"]
        self.assertEqual(updated["B2"].value, "Alex")
        self.assertEqual(updated["B5"].value, "Jamie")
        self.assertEqual(updated["B6"].value, "Morgan")
        self.assertEqual(updated["B8"].value, "Taylor")
        self.assertEqual(updated["B3"].value, "Keep me")
        self.assertEqual(updated["B4"].value, "Keep memo")
        self.assertEqual(updated["B7"].value, "Keep helper")

    def test_reads_existing_primary_assignments(self):
        book = Workbook()
        sheet = book.active
        sheet["B1"] = "9/1"
        sheet["A2"] = "front 10-19"
        sheet["B2"] = "Alex"
        sheet["A3"] = "front help"
        sheet["B3"] = "Ignore helper"
        sheet["A4"] = "Clean A"
        sheet["B4"] = "Jamie"
        sheet["A5"] = "Clean B"
        sheet["A6"] = "宿直(19-10)"
        sheet["B6"] = MANAGER
        source = BytesIO()
        book.save(source)

        result = read_schedule_workbook(
            source, date(2026, 9, 1), date(2026, 9, 30)
        )

        self.assertEqual(
            {(item.position, item.employee) for item in result},
            {("Front", "Alex"), ("Cleaning", "Jamie"), ("Night", MANAGER)},
        )

    def test_reused_template_dates_map_to_selected_year(self):
        book = Workbook()
        sheet = book.active
        sheet["B1"] = date(2025, 9, 1)
        sheet["A2"] = "front 10-19"
        sheet["B2"] = "Alex"
        sheet["A3"] = "Clean A"
        sheet["A4"] = "Clean B"
        sheet["A5"] = "宿直(19-10)"
        sheet["B5"] = MANAGER
        source = BytesIO()
        book.save(source)

        result = read_schedule_workbook(
            source, date(2026, 9, 1), date(2026, 9, 30)
        )

        self.assertEqual({item.day for item in result}, {date(2026, 9, 1)})

    def test_formula_date_is_recognized_for_export(self):
        book = Workbook()
        sheet = book.active
        sheet["B1"] = "=DATE(2025,9,1)"
        sheet["A2"] = "front 10-19"
        sheet["A3"] = "Clean A"
        sheet["A4"] = "Clean B"
        sheet["A5"] = "宿直(19-10)"
        source = BytesIO()
        book.save(source)

        result = update_schedule_workbook(
            source,
            [
                Assignment(date(2026, 9, 1), "Front", "Alex"),
                Assignment(date(2026, 9, 1), "Night", MANAGER),
            ],
            date(2026, 9, 1),
        )

        updated = load_workbook(BytesIO(result), data_only=False).active
        self.assertEqual(updated["B2"].value, "Alex")

    def test_reads_and_updates_generic_schedule_table(self):
        book = Workbook()
        sheet = book.active
        sheet.title = "Schedule"
        sheet.append(["Date", "Position", "Employee", "Source"])
        sheet.append([date(2026, 9, 1), "Front", "Alex", "Edited"])
        source = BytesIO()
        book.save(source)

        preserved = read_schedule_workbook(
            source, date(2026, 9, 1), date(2026, 9, 30)
        )
        self.assertEqual(
            preserved,
            [Assignment(date(2026, 9, 1), "Front", "Alex", "Workbook")],
        )

        result = update_schedule_workbook(
            source,
            [Assignment(date(2026, 9, 1), "Front", "Jamie", "Edited")],
            date(2026, 9, 1),
        )
        updated = load_workbook(BytesIO(result), data_only=True)["Schedule"]
        self.assertEqual(updated["C2"].value, "Jamie")

    def test_creates_missing_october_sheet_from_latest_calendar_layout(self):
        book = Workbook()
        sheet = book.active
        sheet.title = "26.09"
        sheet["A1"] = " "
        for column, weekday in enumerate(
            ("MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"), start=2
        ):
            sheet.cell(1, column).value = weekday
        for date_row in (2, 11, 20, 29, 38, 47):
            sheet.cell(date_row + 1, 1).value = "front 10-19"
            sheet.cell(date_row + 5, 1).value = "Clean A"
            sheet.cell(date_row + 6, 1).value = "Clean B"
            sheet.cell(date_row + 8, 1).value = "宿直(19-10)"
        sheet["B3"] = "Old assignment"
        source = BytesIO()
        book.save(source)

        day = date(2026, 10, 1)
        result = update_schedule_workbook(
            source,
            [
                Assignment(day, "Front", MANAGER),
                Assignment(day, "Cleaning", "Alex"),
                Assignment(day, "Night", MANAGER),
                Assignment(date(2026, 10, 2), "Front", MANAGER),
                Assignment(date(2026, 10, 2), "Cleaning", "Jamie"),
                Assignment(date(2026, 10, 2), "Night", MANAGER),
            ],
            day,
        )

        updated = load_workbook(BytesIO(result), data_only=False)
        self.assertEqual(updated.sheetnames, ["26.09", "26.10"])
        october = updated["26.10"]
        self.assertEqual(october["E2"].value.date(), day)
        self.assertEqual(october["E3"].value, MANAGER)
        self.assertEqual(october["E7"].value, "Alex")
        self.assertEqual(october["E10"].value, MANAGER)
        self.assertEqual(october["F7"].value, "Jamie")
        self.assertIsNone(october["B3"].value)


if __name__ == "__main__":
    unittest.main()

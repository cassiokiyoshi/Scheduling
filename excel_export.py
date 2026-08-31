"""Write generated assignments into the hostel's existing monthly workbook layout."""

from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
import re

from openpyxl import load_workbook

from scheduler import Assignment


def _month_sheet_name(day):
    return f"{day.year % 100:02d}.{day.month:02d}"


def _calendar_date_rows(sheet):
    """Return the date-header rows for each weekly calendar block."""
    rows = []
    for row in range(1, sheet.max_row):
        label = str(sheet.cell(row + 1, 1).value or "").strip().casefold()
        if "front 10-19" in label:
            rows.append(row)
    return rows


def _ensure_month_sheet(workbook, start_day):
    """Create a missing monthly sheet by cloning the latest DEN calendar layout."""
    target_name = _month_sheet_name(start_day)
    if target_name in workbook.sheetnames:
        return workbook[target_name]

    calendar_sheets = [
        sheet for sheet in workbook.worksheets if len(_calendar_date_rows(sheet)) >= 5
    ]
    if not calendar_sheets:
        return None

    source = calendar_sheets[-1]
    sheet = workbook.copy_worksheet(source)
    sheet.title = target_name
    date_rows = _calendar_date_rows(sheet)

    # Clear only the six calendar blocks. Labels, formulas, employee summaries,
    # dimensions, print settings, and visual formatting remain copied intact.
    for date_row in date_rows:
        for row in range(date_row, min(date_row + 9, sheet.max_row + 1)):
            for column in range(2, 9):
                sheet.cell(row, column).value = None

    first_day = date(start_day.year, start_day.month, 1)
    last_day = (
        date(start_day.year + 1, 1, 1) - timedelta(days=1)
        if start_day.month == 12
        else date(start_day.year, start_day.month + 1, 1) - timedelta(days=1)
    )
    current = first_day
    while current <= last_day:
        week_index = (current.day + first_day.weekday() - 1) // 7
        if week_index >= len(date_rows):
            raise ValueError(f"The DEN calendar layout has no room for {target_name}.")
        column = current.weekday() + 2  # Monday=B through Sunday=H.
        sheet.cell(date_rows[week_index], column).value = current
        current += timedelta(days=1)
    return sheet


def _schedule_table_columns(sheet):
    """Return normalized schedule-table columns when the sheet uses that layout."""
    for row in range(1, min(sheet.max_row, 5) + 1):
        columns = {
            str(sheet.cell(row, column).value or "").strip().casefold(): column
            for column in range(1, sheet.max_column + 1)
        }
        if {"date", "position", "employee"}.issubset(columns):
            return row, columns
    return None, None


def _cell_day(value, year, month):
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        if value.month != month:
            return None
        try:
            return date(year, month, value.day)
        except ValueError:
            return None
    if not isinstance(value, str):
        return None
    formula_match = re.search(
        r"DATE\s*\(\s*\d{4}\s*,\s*(\d{1,2})\s*,\s*(\d{1,2})\s*\)",
        value,
        re.IGNORECASE,
    )
    iso_match = re.search(r"\d{4}\s*[-/.]\s*(\d{1,2})\s*[-/.]\s*(\d{1,2})", value)
    short_match = re.search(r"(?<!\d)(\d{1,2})\s*[/.-]\s*(\d{1,2})(?!\d)", value)
    japanese_match = re.search(r"(\d{1,2})\s*月\s*(\d{1,2})\s*日?", value)
    match = formula_match or iso_match or japanese_match or short_match
    if not match or int(match.group(1)) != month:
        return None
    try:
        return date(year, month, int(match.group(2)))
    except ValueError:
        return None


def _row_label(sheet, row, date_column):
    values = [sheet.cell(row, column).value for column in range(1, date_column)]
    return " ".join(str(value).strip() for value in values if value is not None).casefold()


def _position_rows(sheet, date_row, date_column):
    found = {}
    for row in range(date_row + 1, min(date_row + 9, sheet.max_row + 1)):
        label = _row_label(sheet, row, date_column)
        if "front 10-19" in label:
            found["Front"] = row
        elif label.startswith("clean a"):
            found["Cleaning A"] = row
        elif label.startswith("clean b"):
            found["Cleaning B"] = row
        elif "宿直" in label or "19-10" in label:
            found["Night"] = row
    return found


def update_schedule_workbook(workbook_file, assignments, start_day):
    """Return an updated XLSX copy while retaining the workbook's existing styling."""
    workbook_file.seek(0)
    raw_workbook = workbook_file.read()
    workbook = load_workbook(BytesIO(raw_workbook), data_only=False)
    values_workbook = load_workbook(BytesIO(raw_workbook), data_only=True)
    created_sheet = _ensure_month_sheet(workbook, start_day)
    if created_sheet is not None and created_sheet.title not in values_workbook.sheetnames:
        _ensure_month_sheet(values_workbook, start_day)
    by_day = defaultdict(lambda: defaultdict(list))
    for item in assignments:
        by_day[item.day][item.position].append(item.employee)

    # The app's own simple table export can also be uploaded as a preservation
    # source. Update that table directly instead of looking for DEN calendar cells.
    for sheet in workbook.worksheets:
        header_row, columns = _schedule_table_columns(sheet)
        if columns:
            if sheet.max_row > header_row:
                sheet.delete_rows(header_row + 1, sheet.max_row - header_row)
            for row_number, item in enumerate(
                sorted(assignments, key=lambda value: (value.day, value.position, value.employee)),
                start=header_row + 1,
            ):
                sheet.cell(row_number, columns["date"]).value = item.day
                sheet.cell(row_number, columns["position"]).value = item.position
                sheet.cell(row_number, columns["employee"]).value = item.employee
                if "source" in columns:
                    sheet.cell(row_number, columns["source"]).value = item.source
                sheet.cell(row_number, columns["date"]).number_format = "yyyy-mm-dd"
            output = BytesIO()
            workbook.save(output)
            return output.getvalue()

    written_days = set()
    for sheet in workbook.worksheets:
        values_sheet = values_workbook[sheet.title]
        for row in range(1, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):
                displayed_value = values_sheet.cell(row, column).value
                formula_value = sheet.cell(row, column).value
                day = _cell_day(displayed_value, start_day.year, start_day.month)
                if day is None:
                    day = _cell_day(formula_value, start_day.year, start_day.month)
                if day not in by_day:
                    continue
                rows = _position_rows(sheet, row, column)
                if "Front" not in rows or "Night" not in rows:
                    continue
                sheet.cell(rows["Front"], column).value = by_day[day]["Front"][0]
                sheet.cell(rows["Night"], column).value = by_day[day]["Night"][0]
                cleaners = by_day[day]["Cleaning"][:2]
                if "Cleaning A" in rows:
                    sheet.cell(rows["Cleaning A"], column).value = cleaners[0] if cleaners else None
                if "Cleaning B" in rows:
                    sheet.cell(rows["Cleaning B"], column).value = cleaners[1] if len(cleaners) > 1 else None
                written_days.add(day)

    missing = sorted(set(by_day) - written_days)
    if missing:
        raise ValueError(
            "Could not locate calendar cells for: " + ", ".join(day.isoformat() for day in missing)
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def read_schedule_workbook(workbook_file, start_day, end_day):
    """Read existing primary assignments from the DEN calendar workbook."""
    workbook_file.seek(0)
    workbook = load_workbook(workbook_file, data_only=True, read_only=True)
    assignments = []
    seen_days = set()
    for sheet in workbook.worksheets:
        header_row, columns = _schedule_table_columns(sheet)
        if columns:
            for row in range(header_row + 1, sheet.max_row + 1):
                raw_day = sheet.cell(row, columns["date"]).value
                if isinstance(raw_day, datetime):
                    raw_day = raw_day.date()
                elif isinstance(raw_day, str):
                    try:
                        raw_day = date.fromisoformat(raw_day.strip()[:10])
                    except ValueError:
                        continue
                if not isinstance(raw_day, date) or not (start_day <= raw_day <= end_day):
                    continue
                position = str(sheet.cell(row, columns["position"]).value or "").strip()
                employee = str(sheet.cell(row, columns["employee"]).value or "").strip()
                if position in ("Front", "Cleaning", "Night") and employee:
                    assignments.append(Assignment(raw_day, position, employee, "Workbook"))
            return assignments
        for row in range(1, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):
                day = _cell_day(sheet.cell(row, column).value, start_day.year, start_day.month)
                if day is None or not (start_day <= day <= end_day) or day in seen_days:
                    continue
                rows = _position_rows(sheet, row, column)
                if "Front" not in rows or "Night" not in rows:
                    continue
                for position, row_key in (
                    ("Front", "Front"),
                    ("Cleaning", "Cleaning A"),
                    ("Cleaning", "Cleaning B"),
                    ("Night", "Night"),
                ):
                    target_row = rows.get(row_key)
                    value = sheet.cell(target_row, column).value if target_row else None
                    employee = str(value).strip() if value is not None else ""
                    if employee and not employee.startswith("="):
                        assignments.append(Assignment(day, position, employee, "Workbook"))
                seen_days.add(day)
    return assignments
